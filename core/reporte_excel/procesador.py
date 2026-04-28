from pathlib import Path
import re
import pandas as pd
from openpyxl.utils import get_column_letter

FILA_HEADER = 6


def limpiar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).replace("\ufeff", " ").replace("\ufffe", " ")
    texto = re.sub(r"\s+", " ", texto).strip()
    return texto


def buscar_columna(df: pd.DataFrame, nombre_base: str) -> str | None:
    nombre_base_upper = nombre_base.strip().upper()

    for col in df.columns:
        if str(col).strip().upper() == nombre_base_upper:
            return col

    for col in df.columns:
        if str(col).strip().upper().startswith(nombre_base_upper):
            return col

    return None


def buscar_columna_exacta_o_variantes(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    cols_upper = {str(c).strip().upper(): c for c in df.columns}

    for candidato in candidatos:
        candidato_upper = candidato.strip().upper()
        if candidato_upper in cols_upper:
            return cols_upper[candidato_upper]

    for candidato in candidatos:
        candidato_upper = candidato.strip().upper()
        for col in df.columns:
            if str(col).strip().upper().startswith(candidato_upper):
                return col

    return None


def formatear_nombre_apellidos_primero(nombre: str) -> str:
    nombre = limpiar_texto(nombre)
    if not nombre:
        return ""

    partes = nombre.split()
    if len(partes) < 3:
        return nombre

    apellidos = partes[-2:]
    nombres = partes[:-2]
    return f"{' '.join(apellidos)}, {' '.join(nombres)}"


def limpiar_direccion(direccion: str) -> str:
    direccion = limpiar_texto(direccion)
    if not direccion:
        return ""

    patron = r"\b(CALLE|PASAJE|AVENIDA|PJE|AV)\b\.?"
    direccion = re.sub(patron, "", direccion, flags=re.IGNORECASE)
    direccion = re.sub(r"\b(N°|Nº|NO\.?|NRO\.?|NUM\.?)\s*", "", direccion, flags=re.IGNORECASE)
    direccion = direccion.replace("#", " ")
    direccion = re.sub(r"\s+", " ", direccion).strip()
    return direccion


def limpiar_hora(hora: str) -> str:
    hora = limpiar_texto(hora)
    if not hora:
        return ""

    dt = pd.to_datetime(hora, errors="coerce")
    if pd.notna(dt):
        return dt.strftime("%H%M")

    solo_digitos = re.sub(r"\D", "", hora)

    if len(solo_digitos) == 3:
        solo_digitos = "0" + solo_digitos

    if len(solo_digitos) >= 4:
        return solo_digitos[:4]

    return solo_digitos


def transformar_tipo_notificacion(valor: str) -> str:
    valor_limpio = limpiar_texto(valor)

    mapa = {
        "Cedula": "Por cedula",
        "Cédula": "Por cedula",
        "Personal/Art. 44": "Personal/Art44",
        "Personal/Art.44": "Personal/Art44",
        "Personal / Art. 44": "Personal/Art44",
        "Personal / Art.44": "Personal/Art44",
        "Personal/Cedula": "Personal/Cedula",
    }

    return mapa.get(valor_limpio, valor_limpio)


def ajustar_tipo_tramite(tipo_tramite: str, tipo_causa: str) -> str:
    tipo_tramite = limpiar_texto(tipo_tramite)
    tipo_causa = limpiar_texto(tipo_causa).upper()

    if tipo_causa == "E":
        return "Exhorto"

    return tipo_tramite


def normalizar_columna(df: pd.DataFrame, col: str | None, funcion):
    if col is None:
        return pd.Series([""] * len(df), index=df.index)
    return df[col].apply(funcion)


def elegir_gestion_actualizada(c1: str, h1: str, c2: str, h2: str, c3: str, h3: str) -> tuple[str, str]:
    c1, h1 = limpiar_texto(c1), limpiar_texto(h1)
    c2, h2 = limpiar_texto(c2), limpiar_texto(h2)
    c3, h3 = limpiar_texto(c3), limpiar_texto(h3)

    if c3 or h3:
        return c3, h3
    if c2 or h2:
        return c2, h2
    return c1, h1


def aplicar_formato_excel_reporte(df: pd.DataFrame, ruta_salida: Path):
    df = df.copy()

    columnas_numericas = ["RIT", "AÑO", "HORA", "HORA_1", "HORA_2", "HORA_3"]

    for col in columnas_numericas:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    with pd.ExcelWriter(ruta_salida, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Sheet1")
        ws = writer.sheets["Sheet1"]

        columnas_hora = ["HORA", "HORA_1", "HORA_2", "HORA_3"]
        for col_hora in columnas_hora:
            if col_hora in df.columns:
                col_idx = df.columns.get_loc(col_hora) + 1
                for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                    cell = row[col_idx - 1]
                    if cell.value is not None:
                        cell.number_format = "0000"

        for i, col_name in enumerate(df.columns, start=1):
            max_len = len(str(col_name))
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=i, max_col=i):
                val = row[0].value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(i)].width = min(max_len + 2, 45)


def _construir_salida(df: pd.DataFrame, mostrar_todas_gestiones: bool = False) -> pd.DataFrame:
    col_ruc = buscar_columna(df, "RUC")
    col_rit = buscar_columna(df, "RIT")
    col_ano = buscar_columna(df, "AÑO") or buscar_columna(df, "ANO")
    col_nombre = buscar_columna(df, "NOMBRE PARTICIPANTE") or buscar_columna(df, "NOMBRE")
    col_direccion = buscar_columna(df, "DIRECCION") or buscar_columna(df, "DIRECCIÓN")
    col_tipo_tramite = buscar_columna(df, "TIPO TRÁMITE") or buscar_columna(df, "TIPO TRAMITE")
    col_tipo_notificacion = buscar_columna(df, "TIPO NOTIFICACIÓN") or buscar_columna(df, "TIPO NOTIFICACION")
    col_tipo_causa = buscar_columna(df, "TIPO CAUSA")

    col_intento_1 = buscar_columna_exacta_o_variantes(df, ["INTENTO 1", "INTENTO1"])
    col_intento_2 = buscar_columna_exacta_o_variantes(df, ["INTENTO 2", "INTENTO2"])
    col_intento_3 = buscar_columna_exacta_o_variantes(df, ["INTENTO 3", "INTENTO3"])

    col_hora_1 = buscar_columna_exacta_o_variantes(df, ["HORA"])
    col_hora_2 = buscar_columna_exacta_o_variantes(df, ["HORA.1", "HORA 2", "HORA2"])
    col_hora_3 = buscar_columna_exacta_o_variantes(df, ["HORA.2", "HORA 3", "HORA3"])

    faltantes = []
    for etiqueta, col in {
        "RUC": col_ruc,
        "RIT": col_rit,
        "AÑO": col_ano,
        "TIPO TRÁMITE": col_tipo_tramite,
        "TIPO NOTIFICACIÓN": col_tipo_notificacion,
        "NOMBRE PARTICIPANTE": col_nombre,
        "DIRECCION": col_direccion,
        "TIPO CAUSA": col_tipo_causa,
    }.items():
        if col is None:
            faltantes.append(etiqueta)

    if faltantes:
        raise ValueError(f"No se encontraron estas columnas: {', '.join(faltantes)}")

    cod_1 = normalizar_columna(df, col_intento_1, limpiar_texto)
    hora_1 = normalizar_columna(df, col_hora_1, limpiar_hora)

    cod_2 = normalizar_columna(df, col_intento_2, limpiar_texto)
    hora_2 = normalizar_columna(df, col_hora_2, limpiar_hora)

    cod_3 = normalizar_columna(df, col_intento_3, limpiar_texto)
    hora_3 = normalizar_columna(df, col_hora_3, limpiar_hora)

    salida = pd.DataFrame({
        "RUC": df[col_ruc].apply(limpiar_texto),
        "RIT": df[col_rit].apply(limpiar_texto),
        "AÑO": df[col_ano].apply(limpiar_texto),
        "TIPO_TRAMITE": df.apply(
            lambda row: ajustar_tipo_tramite(row[col_tipo_tramite], row[col_tipo_causa]),
            axis=1
        ),
        "TIPO_NOTIFICACION": df[col_tipo_notificacion].apply(transformar_tipo_notificacion),
        "NOMBRE": df[col_nombre].apply(formatear_nombre_apellidos_primero),
        "DIRECCION": df[col_direccion].apply(limpiar_direccion),
    })

    salida[["CODIGO_CERTIFICACION", "HORA"]] = pd.DataFrame(
        [
            elegir_gestion_actualizada(c1, h1, c2, h2, c3, h3)
            for c1, h1, c2, h2, c3, h3 in zip(cod_1, hora_1, cod_2, hora_2, cod_3, hora_3)
        ],
        index=salida.index
    )

    if mostrar_todas_gestiones:
        salida["CODIGO_CERTIFICACION_1"] = cod_1
        salida["HORA_1"] = hora_1
        salida["CODIGO_CERTIFICACION_2"] = cod_2
        salida["HORA_2"] = hora_2
        salida["CODIGO_CERTIFICACION_3"] = cod_3
        salida["HORA_3"] = hora_3

        salida = salida[
            [
                "RUC", "RIT", "AÑO", "TIPO_TRAMITE", "TIPO_NOTIFICACION",
                "NOMBRE", "DIRECCION",
                "CODIGO_CERTIFICACION_1", "HORA_1",
                "CODIGO_CERTIFICACION_2", "HORA_2",
                "CODIGO_CERTIFICACION_3", "HORA_3",
                "CODIGO_CERTIFICACION", "HORA"
            ]
        ]
    else:
        salida = salida[
            [
                "RUC", "RIT", "AÑO", "TIPO_TRAMITE", "TIPO_NOTIFICACION",
                "NOMBRE", "DIRECCION",
                "CODIGO_CERTIFICACION", "HORA"
            ]
        ]

    salida = salida.sort_values(by=["HORA"], ascending=True, na_position="last").reset_index(drop=True)
    return salida


def previsualizar_archivo(ruta_archivo: str | Path, mostrar_todas_gestiones: bool = False, limite: int = 50) -> pd.DataFrame:
    archivo = Path(ruta_archivo)

    if not archivo.exists():
        raise FileNotFoundError(f"No existe el archivo: {archivo}")

    df = pd.read_excel(archivo, header=FILA_HEADER, dtype=str)
    df = df.dropna(how="all").copy()
    df.columns = [limpiar_texto(c) for c in df.columns]

    salida = _construir_salida(df, mostrar_todas_gestiones=mostrar_todas_gestiones)
    return salida.head(limite)


def procesar_archivo(ruta_archivo: str | Path, mostrar_todas_gestiones: bool = False) -> Path:
    archivo = Path(ruta_archivo)

    if not archivo.exists():
        raise FileNotFoundError(f"No existe el archivo: {archivo}")

    df = pd.read_excel(archivo, header=FILA_HEADER, dtype=str)
    df = df.dropna(how="all").copy()
    df.columns = [limpiar_texto(c) for c in df.columns]

    salida = _construir_salida(df, mostrar_todas_gestiones=mostrar_todas_gestiones)

    archivo_salida = archivo.with_name(f"{archivo.stem}_filtrado.xlsx")
    aplicar_formato_excel_reporte(salida, archivo_salida)

    return archivo_salida
