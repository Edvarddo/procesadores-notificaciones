from __future__ import annotations

import re
from datetime import datetime, time
from pathlib import Path
from typing import Dict, List, Any

import xlrd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break


PATRON_SALA = re.compile(r"SALA\s+(\d+)", re.IGNORECASE)
PATRON_ACTA_SALA = re.compile(r"ACTA\s+SALA\s+(\d+)", re.IGNORECASE)
PATRON_PRESENCIAL = re.compile(r"\((PRESENCIAL)\)", re.IGNORECASE)
PATRON_PISO = re.compile(r"(-?\s*\d+°\s*PISO-?)", re.IGNORECASE)
COLUMNAS_TABLA_BASE = ["N°", "RIT", "RUC", "IMPUTADO", "TIPO_AUDIENCIA", "HORA", "TIPO"]

REGISTROS_MAX_POR_PAGINA = 40


def limpiar_texto(valor: object) -> str:
    if valor is None:
        return ""
    texto = str(valor).replace("\n", " ").strip()
    return "" if texto == "nan" else texto


def formatear_fecha(valor: object) -> str:
    return valor.strftime("%d-%m-%Y") if isinstance(valor, datetime) else limpiar_texto(valor)


def formatear_hora(valor: object) -> str:
    if isinstance(valor, (datetime, time)):
        return valor.strftime("%H:%M")
    return limpiar_texto(valor)


def cargar_hoja_plantilla(ruta: Path):
    ext = ruta.suffix.lower()
    if ext == ".xls":
        libro = xlrd.open_workbook(str(ruta))
        for nombre in libro.sheet_names():
            if limpiar_texto(nombre).upper() in {"PLANTILLA", "PLANTILLAS"}:
                return ("xls", libro.sheet_by_name(nombre), nombre)
        raise ValueError("No se encontró una hoja llamada PLANTILLA o PLANTILLAS.")
    libro = load_workbook(ruta, data_only=True)
    for nombre in libro.sheetnames:
        if limpiar_texto(nombre).upper() in {"PLANTILLA", "PLANTILLAS"}:
            return ("xlsx", libro[nombre], nombre)
    raise ValueError("No se encontró una hoja llamada PLANTILLA o PLANTILLAS.")


def obtener_valor(ws, fmt: str, fila_1: int, col_1: int) -> Any:
    if fmt == "xls":
        if fila_1 - 1 >= ws.nrows or col_1 - 1 >= ws.ncols:
            return None
        celda = ws.cell(fila_1 - 1, col_1 - 1)
        valor = celda.value
        if celda.ctype == xlrd.XL_CELL_DATE:
            try:
                return xlrd.xldate_as_datetime(valor, ws.book.datemode)
            except Exception:
                return valor
        if celda.ctype == xlrd.XL_CELL_NUMBER and float(valor).is_integer():
            return int(valor)
        return valor
    return ws.cell(fila_1, col_1).value


def max_filas(ws, fmt: str) -> int:
    return ws.nrows if fmt == "xls" else ws.max_row


def max_cols(ws, fmt: str) -> int:
    return ws.ncols if fmt == "xls" else ws.max_column


def extraer_metadata(ws, fmt: str) -> Dict[str, object]:
    encabezados = {}
    for fila in range(1, min(max_filas(ws, fmt), 10) + 1):
        for col in range(1, min(max_cols(ws, fmt), 12) + 1):
            valor = limpiar_texto(obtener_valor(ws, fmt, fila, col)).upper()
            if valor:
                encabezados[(fila, col)] = valor

    fecha = ""
    tipos_globales = []

    for (fila, col), valor in encabezados.items():
        if valor == "FECHA":
            for f in range(fila + 1, fila + 5):
                candidato = obtener_valor(ws, fmt, f, col)
                if candidato not in (None, ""):
                    if isinstance(candidato, (int, float)) and fmt == "xls":
                        try:
                            candidato = xlrd.xldate_as_datetime(candidato, ws.book.datemode)
                        except Exception:
                            pass
                    fecha = formatear_fecha(candidato)
                    break
        if valor == "TIPO":
            for f in range(fila + 1, fila + 5):
                candidato = limpiar_texto(obtener_valor(ws, fmt, f, col))
                if candidato:
                    tipos_globales.append(candidato)

    encargadas = {1: "", 2: "", 3: "", 4: ""}
    for fila in range(1, 10):
        etiqueta = limpiar_texto(obtener_valor(ws, fmt, fila, 2))
        match = PATRON_ACTA_SALA.search(etiqueta)
        if match:
            encargadas[int(match.group(1))] = limpiar_texto(obtener_valor(ws, fmt, fila, 4))

    return {
        "fecha": fecha,
        "tipo_global": " | ".join(tipos_globales),
        "encargadas": encargadas,
    }


def encontrar_bloques_sala(ws, fmt: str) -> List[Dict[str, Any]]:
    inicios = []
    for fila in range(1, max_filas(ws, fmt) + 1):
        texto = limpiar_texto(obtener_valor(ws, fmt, fila, 1))
        match = PATRON_SALA.search(texto)
        if match:
            sala = int(match.group(1))

            match_presencial = PATRON_PRESENCIAL.search(texto)
            presencial = match_presencial.group(1).upper() if match_presencial else ""

            match_piso = PATRON_PISO.search(texto)
            piso = match_piso.group(1).strip() if match_piso else ""

            inicios.append({
                "sala": sala,
                "inicio": fila,
                "presencial": presencial,
                "piso": piso,
            })

    bloques = []
    for i, item in enumerate(inicios):
        inicio = item["inicio"]
        fin = inicios[i + 1]["inicio"] - 1 if i + 1 < len(inicios) else max_filas(ws, fmt)
        bloques.append({
            "sala": item["sala"],
            "inicio": inicio,
            "fin": fin,
            "presencial": item["presencial"],
            "piso": item["piso"],
        })
    return bloques


def fila_tiene_datos_reales(
    correlativo: str,
    rit: str,
    ruc: str,
    imputados: List[str],
    tipo_audiencia: str,
    hora: str,
    tipo: str,
) -> bool:
    return any(campo for campo in [rit, ruc, tipo_audiencia, hora, tipo, *imputados])


def extraer_registros(ws, fmt: str, metadata: Dict[str, object]) -> List[Dict[str, object]]:
    registros = []
    for bloque in encontrar_bloques_sala(ws, fmt):
        sala = bloque["sala"]
        presencial = bloque.get("presencial", "")
        piso = bloque.get("piso", "")

        for fila in range(bloque["inicio"] + 2, bloque["fin"] + 1):
            correlativo = limpiar_texto(obtener_valor(ws, fmt, fila, 1))
            rit = limpiar_texto(obtener_valor(ws, fmt, fila, 2))
            ruc = limpiar_texto(obtener_valor(ws, fmt, fila, 3))
            imputados = [limpiar_texto(obtener_valor(ws, fmt, fila, col)) for col in range(4, 8)]
            tipo_audiencia = limpiar_texto(obtener_valor(ws, fmt, fila, 9))
            hora = formatear_hora(obtener_valor(ws, fmt, fila, 10))
            tipo = limpiar_texto(obtener_valor(ws, fmt, fila, 11))

            if not fila_tiene_datos_reales(correlativo, rit, ruc, imputados, tipo_audiencia, hora, tipo):
                continue

            imputados_validos = [nombre for nombre in imputados if nombre] or [""]

            for imputado in imputados_validos:
                registros.append({
                    "SALA": sala,
                    "PRESENCIAL": presencial,
                    "PISO": piso,
                    "ENCARGADA_ACTA": metadata["encargadas"].get(sala, ""),
                    "FECHA": metadata["fecha"],
                    "TIPO_GLOBAL": metadata["tipo_global"],
                    "CORRELATIVO": correlativo,
                    "RIT": rit,
                    "RUC": ruc,
                    "IMPUTADO": imputado,
                    "TIPO_AUDIENCIA": tipo_audiencia,
                    "HORA": hora,
                    "TIPO": tipo,
                })
    return registros


THIN_DARK = Side(style="thin", color="7F7F7F")
MEDIUM_DARK = Side(style="medium", color="1F1F1F")
GRIS_GRUPO = PatternFill(fill_type="solid", fgColor="D0D0D0")


def borde_completo(delgado=True):
    side = THIN_DARK if delgado else MEDIUM_DARK
    return Border(left=side, right=side, top=side, bottom=side)


def aplicar_estilo_titulo(ws, fila: int, desde_col: int, hasta_col: int):
    fill = PatternFill("solid", fgColor="1F4E78")
    fuente = Font(bold=True, color="FFFFFF", size=11)
    borde = borde_completo(delgado=False)
    for col in range(desde_col, hasta_col + 1):
        celda = ws.cell(fila, col)
        celda.fill = fill
        celda.font = fuente
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center")


def aplicar_estilo_encabezado(ws, fila: int, desde_col: int, hasta_col: int):
    fill = PatternFill("solid", fgColor="D9EAF7")
    fuente = Font(bold=True, color="000000", size=9)
    borde = borde_completo(delgado=True)
    for col in range(desde_col, hasta_col + 1):
        celda = ws.cell(fila, col)
        celda.fill = fill
        celda.font = fuente
        celda.border = borde
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def aplicar_estilo_metadata(ws, fila: int, cols_etiqueta, cols_valor):
    etiqueta_fill = PatternFill("solid", fgColor="EDEDED")
    bold = Font(bold=True, size=9)
    normal = Font(size=8)
    borde = borde_completo(delgado=True)
    for col in cols_etiqueta:
        c = ws.cell(fila, col)
        c.fill = etiqueta_fill
        c.font = bold
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = borde
    for col in cols_valor:
        c = ws.cell(fila, col)
        c.font = normal
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        c.border = borde


def aplicar_estilo_datos(ws, fila_inicio: int, fila_fin: int, hasta_col: int):
    borde = borde_completo(delgado=True)
    for fila in range(fila_inicio, fila_fin + 1):
        ws.row_dimensions[fila].height = 13
        for col in range(1, hasta_col + 1):
            celda = ws.cell(fila, col)
            celda.alignment = Alignment(
                horizontal="center" if col in {1, 6, 7} else "left",
                vertical="top",
                wrap_text=True,
            )
            celda.border = borde
            celda.font = Font(size=8)


def sombrear_grupos_correlativos(ws, fila_inicio: int, fila_fin: int, hasta_col: int):
    grupos = {}
    for fila in range(fila_inicio, fila_fin + 1):
        correlativo = limpiar_texto(ws.cell(fila, 1).value)
        if correlativo:
            grupos.setdefault(correlativo, []).append(fila)
    for _, filas in grupos.items():
        if len(filas) <= 1:
            continue
        for fila in filas:
            for col in range(1, hasta_col + 1):
                ws.cell(fila, col).fill = GRIS_GRUPO


def ajustar_anchos(ws, incluir_tipo: bool):
    # Base que te gustaba en carta/A4
    anchos = {
        "A": 10,
        "B": 10,
        "C": 10,
        "D": 45,
        "E": 38,
        "F": 8,
    }
    if incluir_tipo:
        anchos["G"] = 15

    # Factor para que en OFICIO horizontal ocupe más hoja
    factor_oficio = 1.27

    for letra, ancho in anchos.items():
        ws.column_dimensions[letra].width = round(ancho * factor_oficio, 2)

def chunked(seq, size):
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def escribir_bloque_sala(
    ws,
    fila_inicio: int,
    sala: int,
    registros_sala: List[Dict[str, object]],
    metadata: Dict[str, object],
    incluir_tipo: bool,
    es_continuacion: bool = False,
) -> int:
    columnas_tabla = COLUMNAS_TABLA_BASE if incluir_tipo else COLUMNAS_TABLA_BASE[:-1]
    ultima_columna = len(columnas_tabla)
    fin_col = get_column_letter(ultima_columna)

    presencial = ""
    piso = ""

    if registros_sala:
        presencial = limpiar_texto(registros_sala[0].get("PRESENCIAL", ""))
        piso = limpiar_texto(registros_sala[0].get("PISO", ""))

    titulo = f"SALA {sala}"
    if presencial:
        titulo += f" ({presencial})"
    if piso:
        titulo += f" - {piso}"
    if es_continuacion:
        titulo += " (CONT.)"

    ws.merge_cells(f"A{fila_inicio}:{fin_col}{fila_inicio}")
    ws.cell(fila_inicio, 1).value = titulo
    aplicar_estilo_titulo(ws, fila_inicio, 1, ultima_columna)
    ws.row_dimensions[fila_inicio].height = 20

    fila_meta = fila_inicio + 1
    ws.cell(fila_meta, 1).value = "Encargada de acta"
    ws.merge_cells(start_row=fila_meta, start_column=2, end_row=fila_meta, end_column=3)
    ws.cell(fila_meta, 2).value = metadata["encargadas"].get(sala, "")
    ws.cell(fila_meta, 4).value = "Fecha"

    if incluir_tipo:
        ws.cell(fila_meta, 5).value = metadata["fecha"]
        ws.cell(fila_meta, 6).value = "Tipo"
        ws.cell(fila_meta, 7).value = metadata["tipo_global"]
        aplicar_estilo_metadata(ws, fila_meta, cols_etiqueta=[1, 4, 6], cols_valor=[2, 3, 5, 7])
    else:
        ws.merge_cells(start_row=fila_meta, start_column=5, end_row=fila_meta, end_column=6)
        ws.cell(fila_meta, 5).value = metadata["fecha"]
        aplicar_estilo_metadata(ws, fila_meta, cols_etiqueta=[1, 4], cols_valor=[2, 3, 5, 6])

    ws.row_dimensions[fila_meta].height = 24

    fila_enc = fila_inicio + 2
    for idx, nombre in enumerate(columnas_tabla, start=1):
        ws.cell(fila_enc, idx).value = nombre
    aplicar_estilo_encabezado(ws, fila_enc, 1, ultima_columna)
    ws.row_dimensions[fila_enc].height = 18

    fila_datos = fila_enc + 1
    if registros_sala:
        for reg in registros_sala:
            ws.cell(fila_datos, 1).value = reg["CORRELATIVO"]
            ws.cell(fila_datos, 2).value = reg["RIT"]
            ws.cell(fila_datos, 3).value = reg["RUC"]
            ws.cell(fila_datos, 4).value = reg["IMPUTADO"]
            ws.cell(fila_datos, 5).value = reg["TIPO_AUDIENCIA"]
            ws.cell(fila_datos, 6).value = reg["HORA"]
            if incluir_tipo:
                ws.cell(fila_datos, 7).value = reg["TIPO"]
            fila_datos += 1
        fila_fin = fila_datos - 1
    else:
        ws.merge_cells(f"A{fila_datos}:{fin_col}{fila_datos}")
        ws.cell(fila_datos, 1).value = "Sin registros"
        ws.cell(fila_datos, 1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(fila_datos, 1).font = Font(size=8)
        ws.row_dimensions[fila_datos].height = 13
        fila_fin = fila_datos

    aplicar_estilo_datos(ws, fila_enc + 1, fila_fin, ultima_columna)
    sombrear_grupos_correlativos(ws, fila_enc + 1, fila_fin, ultima_columna)
    return fila_fin


def escribir_resumen_misma_hoja(ws, metadata: Dict[str, object], incluir_tipo: bool, salas_vacias: List[int]) -> int:
    ultima_col = 7 if incluir_tipo else 6
    fin_col = get_column_letter(ultima_col)

    ws.merge_cells(f"A1:{fin_col}1")
    ws["A1"] = "RESUMEN DE ACTAS"
    aplicar_estilo_titulo(ws, 1, 1, ultima_col)
    ws.row_dimensions[1].height = 22

    ws["A3"] = "SALA"
    ws.merge_cells("B3:D3")
    ws["B3"] = "ENCARGADA DE ACTA"
    ws.merge_cells(f"E3:{fin_col}3")
    ws["E3"] = "FECHA"
    aplicar_estilo_encabezado(ws, 3, 1, ultima_col)

    borde = borde_completo(True)
    fuente = Font(size=8)
    fila = 4
    for sala in range(1, 5):
        ws.cell(fila, 1).value = f"SALA {sala}"
        ws.merge_cells(start_row=fila, start_column=2, end_row=fila, end_column=4)
        ws.cell(fila, 2).value = metadata["encargadas"].get(sala, "")
        ws.merge_cells(start_row=fila, start_column=5, end_row=fila, end_column=ultima_col)
        ws.cell(fila, 5).value = metadata["fecha"]
        for col in range(1, ultima_col + 1):
            c = ws.cell(fila, col)
            c.border = borde
            c.font = fuente
            c.alignment = Alignment(
                horizontal="center" if col in {1, 5} else "left",
                vertical="center",
                wrap_text=True,
            )
        ws.row_dimensions[fila].height = 18
        fila += 1

    for sala in salas_vacias:
        fila += 1
        fila = escribir_bloque_sala(ws, fila, sala, [], metadata, incluir_tipo, es_continuacion=False)

    return fila + 1


def crear_archivo_salida(registros, metadata, ruta_salida: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "AUDIENCIAS"
    incluir_tipo = any(limpiar_texto(r.get("TIPO", "")) for r in registros)

    ajustar_anchos(ws, incluir_tipo)
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_LEGAL
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = False
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.30
    ws.page_margins.right = 0.30
    ws.page_margins.top = 0.40
    ws.page_margins.bottom = 0.40

    salas_con_datos = []
    salas_vacias = []
    for sala in range(1, 5):
        regs_sala = [r for r in registros if r["SALA"] == sala]
        if regs_sala:
            salas_con_datos.append((sala, regs_sala))
        else:
            salas_vacias.append(sala)

    fila_actual = escribir_resumen_misma_hoja(ws, metadata, incluir_tipo, salas_vacias)
    if salas_con_datos:
        ws.row_breaks.append(Break(id=fila_actual - 1))

    for idx_sala, (sala, regs_sala) in enumerate(salas_con_datos):
        paginas = list(chunked(regs_sala, REGISTROS_MAX_POR_PAGINA))
        for idx_pag, regs_pagina in enumerate(paginas):
            fila_fin = escribir_bloque_sala(
                ws,
                fila_actual,
                sala,
                regs_pagina,
                metadata,
                incluir_tipo,
                es_continuacion=(idx_pag > 0),
            )
            fila_actual = fila_fin + 1
            if not (idx_sala == len(salas_con_datos) - 1 and idx_pag == len(paginas) - 1):
                ws.row_breaks.append(Break(id=fila_actual - 1))

    ultima_col = 7 if incluir_tipo else 6
    ws.print_area = f"$A$1:${get_column_letter(ultima_col)}${ws.max_row}"
    wb.save(ruta_salida)


def generar_bitacora(ruta_entrada: str, ruta_salida: str | None = None) -> str:
    ruta_entrada = Path(ruta_entrada)

    if ruta_salida:
        salida = Path(ruta_salida)
    else:
        salida = ruta_entrada.with_name(f"{ruta_entrada.stem}_bitacora.xlsx")

    fmt, ws_origen, _nombre_hoja = cargar_hoja_plantilla(ruta_entrada)
    metadata = extraer_metadata(ws_origen, fmt)
    registros = extraer_registros(ws_origen, fmt, metadata)

    crear_archivo_salida(registros, metadata, salida)
    return str(salida)


def previsualizar_bitacora(ruta_entrada: str):
    ruta_entrada = Path(ruta_entrada)

    fmt, ws_origen, _nombre_hoja = cargar_hoja_plantilla(ruta_entrada)
    metadata = extraer_metadata(ws_origen, fmt)
    registros = extraer_registros(ws_origen, fmt, metadata)

    return registros[:50]
