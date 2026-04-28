from __future__ import annotations

import re
import sys
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook

FILA_HEADER = 6

SLOTS = [
    {"fecha": "E6", "rit_ano": "I6", "nombre": "E8"},
    {"fecha": "E19", "rit_ano": "I19", "nombre": "E21"},
    {"fecha": "E32", "rit_ano": "I32", "nombre": "E34"},
    {"fecha": "E45", "rit_ano": "I45", "nombre": "E47"},
]


def limpiar_texto(valor) -> str:
    if pd.isna(valor):
        return ""
    texto = str(valor).replace("\ufeff", " ").replace("\ufffe", " ")
    return re.sub(r"\s+", " ", texto).strip()


def normalizar_header(valor) -> str:
    texto = limpiar_texto(valor).upper()
    texto = texto.replace("Á", "A").replace("É", "E").replace("Í", "I").replace("Ó", "O").replace("Ú", "U")
    return texto


def buscar_columna_variantes(df: pd.DataFrame, candidatos: list[str]) -> str | None:
    columnas = {normalizar_header(c): c for c in df.columns}

    for candidato in candidatos:
        cand = normalizar_header(candidato)
        if cand in columnas:
            return columnas[cand]

    for candidato in candidatos:
        cand = normalizar_header(candidato)
        for col in df.columns:
            col_norm = normalizar_header(col)
            if col_norm == cand or col_norm.startswith(cand) or cand in col_norm:
                return col

    return None


def detectar_header_row(archivo: Path, max_filas: int = 15) -> int:
    preview = pd.read_excel(archivo, header=None, nrows=max_filas, dtype=str)

    for idx, row in preview.iterrows():
        valores = [normalizar_header(v) for v in row.tolist()]
        joined = " | ".join(v for v in valores if v)

        score = 0
        if "RUC" in joined:
            score += 1
        if "RIT" in joined:
            score += 1
        if "AÑO" in joined or "ANO" in joined:
            score += 1
        if "NOMBRE PARTICIPANTE" in joined or "NOMBRE" in joined:
            score += 1
        if "TIPO NOTIFICACION" in joined:
            score += 1

        if score >= 3:
            return idx

    return 6


def _formatear_nombre(nombre: str) -> str:
    nombre = limpiar_texto(nombre)
    if not nombre:
        return ""
    partes = nombre.split()
    if len(partes) < 3:
        return nombre
    apellidos = partes[-2:]
    nombres = partes[:-2]
    return f"{' '.join(apellidos)}, {' '.join(nombres)}"


def _normalizar_tipo_notificacion(valor: str) -> str:
    v = limpiar_texto(valor).upper()
    v = v.replace(" ", "")
    v = v.replace(".", "")
    v = v.replace("/", "")
    return v


def _es_personal_art44(valor: str) -> bool:
    v = _normalizar_tipo_notificacion(valor)
    return "PERSONAL" in v and "ART44" in v


def _leer_detalle_impresion(ruta_archivo: str | Path) -> pd.DataFrame:
    archivo = Path(ruta_archivo)
    if not archivo.exists():
        raise FileNotFoundError(f"No existe el archivo: {archivo}")

    header_row = detectar_header_row(archivo)
    df = pd.read_excel(archivo, header=header_row, dtype=str)
    df = df.dropna(how="all").copy()
    df.columns = [limpiar_texto(c) for c in df.columns]
    return df


def _construir_salida(df: pd.DataFrame) -> pd.DataFrame:
    col_ruc = buscar_columna_variantes(df, ["RUC"])
    col_rit = buscar_columna_variantes(df, ["RIT"])
    col_ano = buscar_columna_variantes(df, ["AÑO", "ANO"])
    col_nombre = buscar_columna_variantes(df, ["NOMBRE PARTICIPANTE", "NOMBRE"])
    col_tipo_notif = buscar_columna_variantes(df, ["TIPO NOTIFICACIÓN", "TIPO NOTIFICACION"])

    faltantes = []
    for etiqueta, col in {
        "RUC": col_ruc,
        "RIT": col_rit,
        "AÑO": col_ano,
        "NOMBRE": col_nombre,
        "TIPO NOTIFICACIÓN": col_tipo_notif,
    }.items():
        if col is None:
            faltantes.append(etiqueta)

    if faltantes:
        raise ValueError(
            f"No se encontraron estas columnas: {', '.join(faltantes)}. "
            f"Columnas detectadas: {list(df.columns)}"
        )

    df_filtrado = df[df[col_tipo_notif].apply(_es_personal_art44)].copy()

    salida = pd.DataFrame({
        "RUC": df_filtrado[col_ruc].apply(limpiar_texto),
        "RIT": df_filtrado[col_rit].apply(limpiar_texto),
        "AÑO": df_filtrado[col_ano].apply(limpiar_texto),
        "NOMBRE": df_filtrado[col_nombre].apply(_formatear_nombre),
    })

    salida["RIT-AÑO"] = salida.apply(
        lambda row: f"{limpiar_texto(row['RIT'])}-{limpiar_texto(row['AÑO'])}",
        axis=1
    )

    salida = salida[["RUC", "RIT", "AÑO", "NOMBRE", "RIT-AÑO"]].reset_index(drop=True)
    return salida


def previsualizar_avisos(ruta_archivo: str | Path, limite: int | None = 50) -> pd.DataFrame:
    df = _leer_detalle_impresion(ruta_archivo)
    salida = _construir_salida(df)
    return salida if limite is None else salida.head(limite)


def obtener_ano_base(ruta_archivo: str | Path) -> str:
    df = _leer_detalle_impresion(ruta_archivo)
    salida = _construir_salida(df)
    if salida.empty:
        return ""
    anos = salida["AÑO"].astype(str).str.strip()
    anos = anos[anos != ""]
    if anos.empty:
        return ""
    return anos.mode().iloc[0]


def obtener_ruta_asset(nombre_archivo: str) -> Path:
    """
    Devuelve la ruta correcta del asset tanto en desarrollo como en .exe
    empaquetado con PyInstaller.
    """
    if getattr(sys, "frozen", False) and hasattr(sys, "_MEIPASS"):
        base_path = Path(sys._MEIPASS)
    else:
        base_path = Path(__file__).resolve().parents[2]

    return base_path / "assets" / nombre_archivo


def _obtener_ruta_plantilla() -> Path:
    candidatos = [
        obtener_ruta_asset("aviso_notificacion.xlsx"),
        obtener_ruta_asset("Copia de aviso de  Notificacion JG Calama.xlsx"),
    ]

    for ruta in candidatos:
        if ruta.exists():
            return ruta

    raise FileNotFoundError(
        "No se encontro la plantilla de avisos en assets/. "
        "Debe existir como assets/aviso_notificacion.xlsx "
        "o assets/Copia de aviso de  Notificacion JG Calama.xlsx"
    )


def _chunks(seq: list[dict], size: int) -> Iterable[list[dict]]:
    for i in range(0, len(seq), size):
        yield seq[i:i + size]


def _llenar_slot(ws, slot: dict, fecha: str, nombre: str, rit_ano: str):
    ws[slot["fecha"]] = fecha
    ws[slot["rit_ano"]] = rit_ano
    ws[slot["nombre"]] = nombre


def _limpiar_slot(ws, slot: dict):
    ws[slot["fecha"]] = ""
    ws[slot["rit_ano"]] = ""
    ws[slot["nombre"]] = ""


def generar_avisos(ruta_archivo: str | Path, fecha: str, df_final: pd.DataFrame | None = None) -> Path:
    fecha = limpiar_texto(fecha)

    if df_final is None:
        df = _leer_detalle_impresion(ruta_archivo)
        salida = _construir_salida(df)
    else:
        salida = df_final.copy()

    if salida.empty:
        raise ValueError("No hay registros para generar avisos.")

    registros = salida.to_dict(orient="records")

    ruta_plantilla = _obtener_ruta_plantilla()
    wb = load_workbook(ruta_plantilla)
    ws_base = wb["Voucher JGC"]

    grupos = list(_chunks(registros, 4))

    for idx_grupo, grupo in enumerate(grupos):
        if idx_grupo == 0:
            ws = ws_base
            ws.title = "Avisos 1"
        else:
            ws = wb.copy_worksheet(ws_base)
            ws.title = f"Avisos {idx_grupo + 1}"

        for idx_slot, slot in enumerate(SLOTS):
            if idx_slot < len(grupo):
                reg = grupo[idx_slot]
                _llenar_slot(
                    ws,
                    slot,
                    fecha=fecha,  # puede ir vacía
                    nombre=limpiar_texto(reg["NOMBRE"]),
                    rit_ano=limpiar_texto(reg["RIT-AÑO"]),
                )
            else:
                _limpiar_slot(ws, slot)

    if len(grupos) == 1:
        ws_base.title = "Avisos 1"

    archivo = Path(ruta_archivo)
    salida_path = archivo.with_name(f"{archivo.stem}_avisos.xlsx")
    wb.save(salida_path)
    return salida_path
